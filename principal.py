from ast import For
import openpyxl 
  
#path = "C:\\proyectos\\boletomovil-chivas-vs-puebla.xlsx"
path = "C:\\Users\\hersc\\OneDrive\\Escritorio\\correos\\monedero.xlsx"
path2 = "C:\\Users\\hersc\\OneDrive\\Escritorio\\correos\\monedero2.xlsx"

  
wb_obj = openpyxl.load_workbook(path) 
  
sheet_obj = wb_obj.active 
    
cell_obj = sheet_obj.cell(row = 1, column = 1) 


#toma un correo del path 1 y pregunta si existe en el path2
#si no existe lo agrega al path2 y limpia el registro en el path1

wb_obj2 = openpyxl.load_workbook(path2) 
  
sheet_obj2 = wb_obj2.active 
    
cell_obj2 = sheet_obj2.cell(row = 1, column = 1) 



correos = []

for i in range(1, sheet_obj.max_row):

    correo = str(sheet_obj.cell(row = i, column = 1).value) 
    columa2 = str(sheet_obj.cell(row = i, column = 2).value)
    columa3 = str(sheet_obj.cell(row = i, column = 3).value)
    columa4 = str(sheet_obj.cell(row = i, column = 4).value)
    columa5 = str(sheet_obj.cell(row = i, column = 5).value)

    existe = False

    for j in range(1,sheet_obj2.max_row+1):
        correoAux = str(sheet_obj2.cell(row = j, column = 1).value)
        if(correo==correoAux):
            existe=True
    
    if(existe==False):
        tope = sheet_obj2.max_row+1
        sheet_obj2.cell(row = tope, column = 1).value=correo
        sheet_obj2.cell(row = tope, column = 2).value=columa2
        sheet_obj2.cell(row = tope, column = 3).value=columa3
        sheet_obj2.cell(row = tope, column = 4).value=columa4
        sheet_obj2.cell(row = tope, column = 5).value=columa5
        sheet_obj.cell(row = i, column = 1).value=""
        sheet_obj.cell(row = i, column = 2).value=""
        sheet_obj.cell(row = i, column = 3).value=""
        sheet_obj.cell(row = i, column = 4).value=""
        sheet_obj.cell(row = i, column = 5).value=""
        wb_obj2.save("C:\\Users\\hersc\\OneDrive\\Escritorio\\correos\\monedero2.xlsx")


   
wb_obj.save("C:\\Users\\hersc\\OneDrive\\Escritorio\\correos\\monedero.xlsx")
wb_obj2.save("C:\\Users\\hersc\\OneDrive\\Escritorio\\correos\\monedero2.xlsx")